import csv
import io
import json
from functools import wraps

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import ensure_csrf_cookie
from django.db.models import Q, Max

from reconnect.models import (
    CustomUser, Conversation, ConversationParticipant, Message,
    Event, EventTimelineItem, Announcement,
    Post, PostLike, PostComment, Connection, Opportunity, Project,
)


# ─── Role decorator ──────────────────────────────────────────────────────────

def role_required(*roles):
    """Restrict view to users with the specified role(s)."""
    def decorator(view_func):
        @wraps(view_func)
        @login_required
        def wrapper(request, *args, **kwargs):
            if request.user.role not in roles:
                return JsonResponse({'error': 'Forbidden'}, status=403)
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


# ─── Auth & Page Views ────────────────────────────────────────────────────────

def home(request):
    return render(request, "home.html")


def login_view(request):
    if request.method == "POST":
        enrollment = request.POST.get("enrollment_number")
        password = request.POST.get("password")
        user = authenticate(request, username=enrollment, password=password)

        if user is not None:
            login(request, user)
            if user.role == "admin" or user.is_superuser:
                return redirect("admin_dashboard")
            elif user.role == "student":
                return redirect("student_dashboard")
            elif user.role == "alumni":
                return redirect("alumni_dashboard")
            else:
                return redirect("home")
        else:
            messages.error(request, "Invalid enrollment number or password")

    return render(request, "login.html")


def logout_view(request):
    logout(request)
    return redirect("login")


@login_required
@ensure_csrf_cookie
def admin_dashboard(request):
    events_list = Event.objects.filter(is_active=True)
    announcements_list = Announcement.objects.filter(is_active=True)
    return render(request, "admin.html", {
        'events': events_list,
        'announcements': announcements_list,
    })


@login_required
def student_dashboard(request):
    events_list = Event.objects.filter(is_active=True)
    connection_count = Connection.objects.filter(
        Q(from_user=request.user, status='accepted') | Q(to_user=request.user, status='accepted')
    ).count()
    post_count = Post.objects.filter(author=request.user).count()
    return render(request, "student/studentdash.html", {
        'user': request.user,
        'events': events_list,
        'connection_count': connection_count,
        'post_count': post_count,
    })


@login_required
def alumni_dashboard(request):
    events_list = Event.objects.filter(is_active=True)
    announcements_list = Announcement.objects.filter(is_active=True)
    return render(request, "alumini/aluminidash.html", {
        'user': request.user,
        'events': events_list,
        'announcements': announcements_list,
    })


@login_required
def events(request):
    events_list = Event.objects.filter(is_active=True)
    return render(request, "alumini/events.html", {'user': request.user, 'events': events_list})


@login_required
def announcements(request):
    announcements_list = Announcement.objects.filter(is_active=True)
    return render(request, "alumini/announcements.html", {'user': request.user, 'announcements': announcements_list})


@login_required
def connect(request):
    return render(request, "alumini/connect.html", {'user': request.user})


@login_required
def settings_view(request):
    return render(request, "alumini/settings.html", {'user': request.user})


@login_required
def profile(request):
    # If ?id= is provided, show that user's profile; otherwise show own
    user_id = request.GET.get('id')
    if user_id:
        profile_user = get_object_or_404(CustomUser, id=user_id)
    else:
        enr = request.GET.get('enr')
        if enr:
            profile_user = get_object_or_404(CustomUser, enrollment_number=enr)
        else:
            profile_user = request.user
    return render(request, "alumini/profile.html", {'profile_user': profile_user})


@login_required
def event_details(request):
    event_id = request.GET.get('id')
    event = None
    timeline = []
    events_list = Event.objects.filter(is_active=True)
    if event_id:
        event = Event.objects.filter(id=event_id, is_active=True).first()
        if event:
            timeline = event.timeline.all()
    template = 'student/eventdetails2.html' if request.user.role == 'student' else 'alumini/eventdetails.html'
    return render(request, template, {'event': event, 'timeline': timeline, 'events': events_list, 'user': request.user})


# ─── Student Page Views ──────────────────────────────────────────────────────

@login_required
def student_connect(request):
    return render(request, "student/student_connect.html", {'user': request.user})


@login_required
def student_explore(request):
    return render(request, "student/student_explore.html", {'user': request.user})


@login_required
def student_opportunities(request):
    opportunities = Opportunity.objects.filter(is_active=True)
    return render(request, "student/student_opportunities.html", {
        'user': request.user,
        'opportunities': opportunities,
    })


@login_required
def student_projects(request):
    projects = Project.objects.filter(is_active=True)
    return render(request, "student/student_projects.html", {
        'user': request.user,
        'projects': projects,
    })


@login_required
def student_profile(request):
    connection_count = Connection.objects.filter(
        Q(from_user=request.user, status='accepted') | Q(to_user=request.user, status='accepted')
    ).count()
    post_count = Post.objects.filter(author=request.user).count()
    project_count = Project.objects.filter(posted_by=request.user).count()
    return render(request, "student/student_profile.html", {
        'user': request.user,
        'connection_count': connection_count,
        'post_count': post_count,
        'project_count': project_count,
    })


@login_required
def student_settings(request):
    return render(request, "student/student_settings.html", {'user': request.user})


# ─── Alumni Extra Page Views ─────────────────────────────────────────────────

@login_required
def my_profile(request):
    return render(request, "alumini/myprofile.html", {'user': request.user})


@login_required
def post_view(request):
    post_count = Post.objects.filter(author=request.user).count()
    grant_count = Post.objects.filter(author=request.user, post_type='funding').count()
    return render(request, "alumini/post.html", {
        'user': request.user,
        'post_count': post_count,
        'grant_count': grant_count,
    })


# ─── Post / Social Feed API ──────────────────────────────────────────────────

@require_POST
@login_required
def create_post(request):
    """Create a social post (general, hiring, funding, openfor)."""
    post_type = request.POST.get('post_type', 'general').strip()
    title = request.POST.get('title', '').strip()
    body = request.POST.get('body', '').strip()

    post = Post(
        author=request.user,
        post_type=post_type,
        title=title,
        body=body,
    )

    # Handle image upload
    img = request.FILES.get('image')
    if img:
        post.image = img

    # Hiring-specific fields
    if post_type == 'hiring':
        post.company = request.POST.get('company', '').strip()
        post.role = request.POST.get('role', '').strip()
        post.job_type = request.POST.get('job_type', '').strip()
        post.duration = request.POST.get('duration', '').strip()
        post.stipend = request.POST.get('stipend', '').strip()
        post.location = request.POST.get('location', '').strip()
        post.application_url = request.POST.get('application_url', '').strip()

    # Funding-specific
    if post_type == 'funding':
        post.amount = request.POST.get('amount', '').strip()
        post.frequency = request.POST.get('frequency', '').strip()
        post.eligibility = request.POST.get('eligibility', '').strip()

    # Open-for
    if post_type == 'openfor':
        tags = request.POST.getlist('open_for_tags')
        post.open_for_tags = ','.join(tags)

    post.save()
    return JsonResponse({'success': True, 'message': 'Post published!', 'id': post.id})


@require_GET
@login_required
def api_posts_list(request):
    """Return feed posts."""
    qs = Post.objects.filter(is_active=True).select_related('author')
    # Optional filter by user
    user_id = request.GET.get('user_id')
    if user_id:
        qs = qs.filter(author_id=user_id)
    posts = qs[:50]
    user = request.user
    result = []
    for p in posts:
        result.append({
            'id': p.id,
            'post_type': p.post_type,
            'title': p.title,
            'body': p.body,
            'image': p.image.url if p.image else '',
            'author_name': p.author.get_full_name() or p.author.username,
            'author_initials': p.author.get_initials(),
            'author_department': p.author.department,
            'author_id': p.author.id,
            'author_profile_picture': p.author.profile_picture.url if p.author.profile_picture else '',
            'company': p.company,
            'role': p.role,
            'job_type': p.job_type,
            'location': p.location,
            'amount': p.amount,
            'open_for_tags': p.open_for_tags.split(',') if p.open_for_tags else [],
            'likes': p.like_count(),
            'comments': p.comment_count(),
            'liked_by_me': p.likes.filter(user=user).exists(),
            'created_at': p.created_at.strftime('%b %d, %Y %H:%M'),
        })
    return JsonResponse({'posts': result})


@require_POST
@login_required
def toggle_like(request, post_id):
    """Toggle like on a post."""
    post = get_object_or_404(Post, id=post_id)
    like, created = PostLike.objects.get_or_create(post=post, user=request.user)
    if not created:
        like.delete()
        return JsonResponse({'liked': False, 'count': post.like_count()})
    return JsonResponse({'liked': True, 'count': post.like_count()})


@require_POST
@login_required
def add_comment(request, post_id):
    """Add a comment to a post."""
    post = get_object_or_404(Post, id=post_id)
    try:
        data = json.loads(request.body)
        content = data.get('content', '').strip()
    except json.JSONDecodeError:
        content = request.POST.get('content', '').strip()
    if not content:
        return JsonResponse({'error': 'Comment cannot be empty'}, status=400)
    comment = PostComment.objects.create(post=post, user=request.user, content=content)
    return JsonResponse({
        'success': True,
        'comment': {
            'id': comment.id,
            'content': comment.content,
            'user_name': comment.user.get_full_name() or comment.user.username,
            'user_initials': comment.user.get_initials(),
            'created_at': comment.created_at.strftime('%b %d, %Y %H:%M'),
        }
    })


@require_GET
@login_required
def get_comments(request, post_id):
    """Get comments for a post."""
    post = get_object_or_404(Post, id=post_id)
    comments = post.comments.select_related('user').all()
    result = [{
        'id': c.id,
        'content': c.content,
        'user_name': c.user.get_full_name() or c.user.username,
        'user_initials': c.user.get_initials(),
        'created_at': c.created_at.strftime('%b %d, %Y %H:%M'),
    } for c in comments]
    return JsonResponse({'comments': result})


# ─── Connection API ───────────────────────────────────────────────────────────

@require_POST
@login_required
def send_connection_request(request):
    """Send a connection request."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    to_user_id = data.get('to_user_id')
    if not to_user_id:
        return JsonResponse({'error': 'to_user_id required'}, status=400)

    try:
        to_user = CustomUser.objects.get(id=to_user_id)
    except CustomUser.DoesNotExist:
        return JsonResponse({'error': 'User not found'}, status=404)

    if to_user == request.user:
        return JsonResponse({'error': 'Cannot connect with yourself'}, status=400)

    conn, created = Connection.objects.get_or_create(
        from_user=request.user, to_user=to_user,
        defaults={'status': 'pending'}
    )
    if not created:
        return JsonResponse({'status': conn.status, 'message': 'Request already exists'})
    return JsonResponse({'success': True, 'status': 'pending', 'message': 'Connection request sent'})


@require_POST
@login_required
def respond_connection(request, connection_id):
    """Accept or decline a connection request."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    action = data.get('action')  # 'accept' or 'decline'
    conn = get_object_or_404(Connection, id=connection_id, to_user=request.user)

    if action == 'accept':
        conn.status = 'accepted'
        conn.save()
        return JsonResponse({'success': True, 'status': 'accepted'})
    elif action == 'decline':
        conn.status = 'declined'
        conn.save()
        return JsonResponse({'success': True, 'status': 'declined'})
    return JsonResponse({'error': 'Invalid action'}, status=400)


@require_GET
@login_required
def connection_list(request):
    """Get user's connections and pending requests."""
    user = request.user
    # Accepted connections
    accepted = Connection.objects.filter(
        Q(from_user=user, status='accepted') | Q(to_user=user, status='accepted')
    ).select_related('from_user', 'to_user')

    connections = []
    for c in accepted:
        other = c.to_user if c.from_user == user else c.from_user
        connections.append({
            'id': c.id,
            'user_id': other.id,
            'name': other.get_full_name() or other.username,
            'initials': other.get_initials(),
            'department': other.department,
            'enrollment_number': other.enrollment_number,
            'profile_picture': other.profile_picture.url if other.profile_picture else '',
        })

    # Pending requests received
    pending = Connection.objects.filter(to_user=user, status='pending').select_related('from_user')
    requests_list = [{
        'id': c.id,
        'user_id': c.from_user.id,
        'name': c.from_user.get_full_name() or c.from_user.username,
        'initials': c.from_user.get_initials(),
        'department': c.from_user.department,
        'profile_picture': c.from_user.profile_picture.url if c.from_user.profile_picture else '',
    } for c in pending]

    return JsonResponse({'connections': connections, 'pending_requests': requests_list})


# ─── Opportunity & Project API ────────────────────────────────────────────────

@require_GET
@login_required
def api_opportunities_list(request):
    """Return active opportunities + hiring posts."""
    # From Opportunity model
    opps = Opportunity.objects.filter(is_active=True)
    result = [{
        'id': o.id,
        'title': o.title,
        'company': o.company,
        'type': o.opportunity_type,
        'description': o.description,
        'location': o.location,
        'stipend': o.stipend,
        'application_url': o.application_url,
        'posted_by': o.posted_by.get_full_name() if o.posted_by else '',
        'poster_profile_picture': o.posted_by.profile_picture.url if (o.posted_by and o.posted_by.profile_picture) else '',
        'created_at': o.created_at.strftime('%b %d, %Y'),
    } for o in opps]

    # Also include hiring posts from Post model
    hiring_posts = Post.objects.filter(is_active=True, post_type='hiring').select_related('author')
    for p in hiring_posts:
        result.append({
            'id': f'post-{p.id}',
            'title': p.title or p.role or 'Hiring',
            'company': p.company or '',
            'type': 'internship' if p.job_type and 'intern' in p.job_type.lower() else 'fulltime',
            'description': p.body,
            'location': p.location or 'Remote',
            'stipend': p.stipend or 'Not specified',
            'application_url': p.application_url or '',
            'posted_by': p.author.get_full_name() or p.author.username,
            'poster_profile_picture': p.author.profile_picture.url if p.author.profile_picture else '',
            'created_at': p.created_at.strftime('%b %d, %Y'),
        })

    return JsonResponse({'opportunities': result})


@require_GET
@login_required
def api_projects_list(request):
    """Return active projects + funding/grant posts."""
    # From Project model
    projs = Project.objects.filter(is_active=True)
    result = [{
        'id': p.id,
        'title': p.title,
        'category': p.category,
        'description': p.description,
        'tech_stack': p.tech_stack,
        'team_size': p.team_size,
        'posted_by': p.posted_by.get_full_name() if p.posted_by else '',
        'poster_profile_picture': p.posted_by.profile_picture.url if (p.posted_by and p.posted_by.profile_picture) else '',
        'created_at': p.created_at.strftime('%b %d, %Y'),
    } for p in projs]

    # Also include funding/grant posts from Post model
    funding_posts = Post.objects.filter(is_active=True, post_type='funding').select_related('author')
    for fp in funding_posts:
        result.append({
            'id': f'post-{fp.id}',
            'title': fp.title or 'Grant / Funding',
            'category': 'research',
            'description': fp.body,
            'tech_stack': '',
            'team_size': 1,
            'posted_by': fp.author.get_full_name() or fp.author.username,
            'poster_profile_picture': fp.author.profile_picture.url if fp.author.profile_picture else '',
            'created_at': fp.created_at.strftime('%b %d, %Y'),
            'amount': fp.amount or '',
            'eligibility': fp.eligibility or '',
        })

    return JsonResponse({'projects': result})


# ─── Explore / People Search API ─────────────────────────────────────────────

@require_GET
@login_required
def api_explore_people(request):
    """Search for people with optional role filter."""
    q = request.GET.get('q', '').strip()
    role_filter = request.GET.get('role', '').strip()

    qs = CustomUser.objects.exclude(id=request.user.id)

    if q and len(q) >= 2:
        qs = qs.filter(
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(enrollment_number__icontains=q) |
            Q(department__icontains=q)
        )

    if role_filter and role_filter != 'all':
        qs = qs.filter(role=role_filter)

    users = qs[:30]
    user = request.user

    result = []
    for u in users:
        # Check connection status
        conn = Connection.objects.filter(
            Q(from_user=user, to_user=u) | Q(from_user=u, to_user=user)
        ).first()
        conn_status = conn.status if conn else 'none'

        result.append({
            'id': u.id,
            'name': u.get_full_name() or u.username,
            'initials': u.get_initials(),
            'department': u.department,
            'role': u.role,
            'enrollment_number': u.enrollment_number,
            'passed_out_year': u.passed_out_year,
            'working_status': u.working_status,
            'connection_status': conn_status,
            'profile_picture': u.profile_picture.url if u.profile_picture else '',
        })
    return JsonResponse({'people': result})


# ─── CSV / Bulk Upload ───────────────────────────────────────────────────────

@require_POST
@login_required
def bulk_upload_users(request):
    """
    Accept CSV/XLSX file upload.  Creates CustomUser records.
    Expects a 'default_password' field and a 'role' field in the POST body.
    CSV columns: Enrollment No, Name, Department, Passed Year, CGPA, Status, Phone
    """
    uploaded_file = request.FILES.get('file')
    default_password = request.POST.get('default_password', 'connect@123')
    default_role = request.POST.get('role', 'alumni')

    if not uploaded_file:
        return JsonResponse({'error': 'No file uploaded'}, status=400)

    created = 0
    skipped = 0
    errors = []

    try:
        filename = uploaded_file.name.lower()

        if filename.endswith('.csv'):
            decoded = uploaded_file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(decoded))
            rows = list(reader)
        elif filename.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(uploaded_file, read_only=True)
            ws = wb.active
            headers = [str(cell.value or '').strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, [str(v) if v is not None else '' for v in row])))
        else:
            return JsonResponse({'error': 'Unsupported file format. Use .csv or .xlsx'}, status=400)

        for i, row in enumerate(rows, start=2):
            try:
                enrollment = (row.get('Enrollment No') or row.get('enrollment_no') or '').strip()
                name = (row.get('Name') or row.get('name') or '').strip()
                department = (row.get('Department') or row.get('department') or '').strip()
                passed_year = (row.get('Passed Year') or row.get('passed_year') or '').strip()
                cgpa = (row.get('CGPA') or row.get('cgpa') or '').strip()
                status = (row.get('Status') or row.get('status') or '').strip()
                phone = (row.get('Phone') or row.get('phone') or '').strip()

                if not enrollment:
                    errors.append(f"Row {i}: Missing enrollment number")
                    skipped += 1
                    continue

                if CustomUser.objects.filter(enrollment_number=enrollment).exists():
                    errors.append(f"Row {i}: {enrollment} already exists")
                    skipped += 1
                    continue

                # Split name
                parts = name.split(maxsplit=1)
                first_name = parts[0] if parts else ''
                last_name = parts[1] if len(parts) > 1 else ''

                user = CustomUser(
                    enrollment_number=enrollment,
                    username=enrollment,  # use enrollment as username
                    first_name=first_name,
                    last_name=last_name,
                    department=department,
                    passed_out_year=int(passed_year) if passed_year else None,
                    cgpa=float(cgpa) if cgpa else None,
                    working_status=status,
                    phone=phone,
                    role=default_role,
                )
                user.set_password(default_password)
                user.save()
                created += 1

            except Exception as e:
                errors.append(f"Row {i}: {str(e)}")
                skipped += 1

    except Exception as e:
        return JsonResponse({'error': f'File processing error: {str(e)}'}, status=500)

    return JsonResponse({
        'created': created,
        'skipped': skipped,
        'errors': errors,
    })


@require_POST
@login_required
def create_single_user(request):
    """Create a single user from the admin form."""
    try:
        data = request.POST
        enrollment = data.get('enrollment_number', '').strip()
        name = data.get('full_name', '').strip()
        department = data.get('department', '').strip()
        passed_year = data.get('passed_out_year', '').strip()
        cgpa = data.get('cgpa', '').strip()
        status = data.get('working_status', '').strip()
        phone = data.get('phone', '').strip()
        role = data.get('role', 'alumni').strip()
        password = data.get('password', 'connect@123').strip()

        if not enrollment:
            return JsonResponse({'error': 'Enrollment number is required'}, status=400)

        if CustomUser.objects.filter(enrollment_number=enrollment).exists():
            return JsonResponse({'error': 'Enrollment number already exists'}, status=400)

        parts = name.split(maxsplit=1)
        first_name = parts[0] if parts else ''
        last_name = parts[1] if len(parts) > 1 else ''

        user = CustomUser(
            enrollment_number=enrollment,
            username=enrollment,
            first_name=first_name,
            last_name=last_name,
            department=department,
            passed_out_year=int(passed_year) if passed_year else None,
            cgpa=float(cgpa) if cgpa else None,
            working_status=status,
            phone=phone,
            role=role,
        )
        user.set_password(password)
        user.save()

        return JsonResponse({'success': True, 'message': f'User {enrollment} created successfully'})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ─── Profile / Settings API ──────────────────────────────────────────────────

@require_POST
@login_required
def update_profile(request):
    """Handle profile settings update: phone, working_status, password, profile picture."""
    user = request.user
    updated = []

    # Phone update
    phone = request.POST.get('phone')
    if phone is not None:
        user.phone = phone.strip()
        updated.append('phone')

    # Working status update
    ws = request.POST.get('working_status')
    if ws is not None:
        user.working_status = ws.strip()
        updated.append('working_status')

    # Profile picture upload
    pic = request.FILES.get('profile_picture')
    if pic:
        user.profile_picture = pic
        updated.append('profile_picture')

    # Password change
    current_pw = request.POST.get('current_password', '').strip()
    new_pw = request.POST.get('new_password', '').strip()
    confirm_pw = request.POST.get('confirm_password', '').strip()

    if new_pw:
        if not current_pw:
            return JsonResponse({'error': 'Current password is required'}, status=400)
        if not user.check_password(current_pw):
            return JsonResponse({'error': 'Current password is incorrect'}, status=400)
        if new_pw != confirm_pw:
            return JsonResponse({'error': 'New passwords do not match'}, status=400)
        if len(new_pw) < 6:
            return JsonResponse({'error': 'Password must be at least 6 characters'}, status=400)
        user.set_password(new_pw)
        updated.append('password')

    user.save()

    # Keep user logged in after password change
    if 'password' in updated:
        update_session_auth_hash(request, user)

    return JsonResponse({
        'success': True,
        'message': f'Updated: {", ".join(updated)}' if updated else 'No changes made',
        'profile_picture_url': user.profile_picture.url if user.profile_picture else '',
    })


# ─── Chat API ────────────────────────────────────────────────────────────────

@require_GET
@login_required
def conversation_list(request):
    """Return the current user's conversations with last message info."""
    user = request.user
    convos = Conversation.objects.filter(
        membership__user=user
    ).annotate(
        last_msg_time=Max('messages__timestamp')
    ).order_by('-last_msg_time')

    result = []
    for c in convos:
        last_msg = c.last_message()
        # For 1-on-1, get the other person's name
        if not c.is_group:
            other = c.membership.exclude(user=user).select_related('user').first()
            display_name = other.user.get_full_name() or other.user.username if other else 'Unknown'
            initials = other.user.get_initials() if other else '??'
        else:
            display_name = c.name or 'Group'
            initials = 'GRP'

        result.append({
            'id': str(c.id),
            'name': display_name,
            'initials': initials,
            'is_group': c.is_group,
            'profile_picture': other.user.profile_picture.url if (not c.is_group and other and other.user.profile_picture) else '',
            'last_message': last_msg.content[:50] if last_msg else '',
            'last_message_sender': (last_msg.sender.first_name or last_msg.sender.username) if last_msg else '',
            'last_message_time': last_msg.timestamp.strftime('%H:%M') if last_msg else '',
        })

    return JsonResponse({'conversations': result})


@require_POST
@login_required
def conversation_create(request):
    """
    Create a new conversation.
    For 1-on-1: POST { user_id: <id> }
    For group : POST { name: '...', user_ids: [id1, id2, ...], is_group: true }
    """
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    is_group = data.get('is_group', False)
    user = request.user

    if is_group:
        name = data.get('name', 'New Group')
        user_ids = data.get('user_ids', [])

        convo = Conversation.objects.create(
            name=name,
            is_group=True,
            created_by=user,
        )
        ConversationParticipant.objects.create(conversation=convo, user=user)
        for uid in user_ids:
            try:
                member = CustomUser.objects.get(id=uid)
                ConversationParticipant.objects.get_or_create(conversation=convo, user=member)
            except CustomUser.DoesNotExist:
                pass

        return JsonResponse({'id': str(convo.id), 'name': convo.name})

    else:
        other_id = data.get('user_id')
        if not other_id:
            return JsonResponse({'error': 'user_id is required'}, status=400)

        try:
            other_user = CustomUser.objects.get(id=other_id)
        except CustomUser.DoesNotExist:
            return JsonResponse({'error': 'User not found'}, status=404)

        # Check for existing 1-on-1 conversation
        existing = Conversation.objects.filter(
            is_group=False,
            membership__user=user,
        ).filter(
            membership__user=other_user,
        ).first()

        if existing:
            return JsonResponse({'id': str(existing.id), 'name': other_user.get_full_name(), 'existing': True})

        convo = Conversation.objects.create(is_group=False, created_by=user)
        ConversationParticipant.objects.create(conversation=convo, user=user)
        ConversationParticipant.objects.create(conversation=convo, user=other_user)

        return JsonResponse({
            'id': str(convo.id),
            'name': other_user.get_full_name() or other_user.username,
        })


@require_GET
@login_required
def conversation_messages(request, conversation_id):
    """Return message history for a conversation (paginated)."""
    user = request.user

    # Ensure user is a participant
    if not ConversationParticipant.objects.filter(conversation_id=conversation_id, user=user).exists():
        return JsonResponse({'error': 'Not a participant'}, status=403)

    page = int(request.GET.get('page', 1))
    page_size = 50
    offset = (page - 1) * page_size

    msgs = Message.objects.filter(
        conversation_id=conversation_id
    ).select_related('sender').order_by('-timestamp')[offset:offset + page_size]

    result = [{
        'id': str(m.id),
        'content': m.content,
        'sender_id': m.sender.id,
        'sender_name': m.sender.get_full_name() or m.sender.username,
        'sender_initials': m.sender.get_initials(),
        'timestamp': m.timestamp.strftime('%H:%M'),
        'is_mine': m.sender.id == user.id,
    } for m in reversed(msgs)]

    return JsonResponse({'messages': result, 'page': page})


@require_POST
@login_required
def send_message(request, conversation_id):
    """Send a message to a conversation via HTTP (fallback when WebSocket is unavailable)."""
    user = request.user

    if not ConversationParticipant.objects.filter(conversation_id=conversation_id, user=user).exists():
        return JsonResponse({'error': 'Not a participant'}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    content = data.get('content', '').strip()
    if not content:
        return JsonResponse({'error': 'Empty message'}, status=400)

    conversation = get_object_or_404(Conversation, id=conversation_id)
    msg = Message.objects.create(conversation=conversation, sender=user, content=content)

    return JsonResponse({
        'id': str(msg.id),
        'content': msg.content,
        'sender_id': user.id,
        'timestamp': msg.timestamp.strftime('%H:%M'),
        'success': True,
    })


@require_GET
@login_required
def user_search(request):
    """Search users by name or enrollment number for new chat / group creation."""
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'users': []})

    users = CustomUser.objects.filter(
        Q(first_name__icontains=q) |
        Q(last_name__icontains=q) |
        Q(enrollment_number__icontains=q) |
        Q(username__icontains=q)
    ).exclude(id=request.user.id)[:20]

    result = [{
        'id': u.id,
        'name': u.get_full_name() or u.username,
        'enrollment_number': u.enrollment_number,
        'department': u.department,
        'initials': u.get_initials(),
    } for u in users]

    return JsonResponse({'users': result})


# ─── Event & Announcement API ────────────────────────────────────────────────

@require_POST
@login_required
def create_event(request):
    """Create an event from the admin panel."""
    title = request.POST.get('title', '').strip()
    date_display = request.POST.get('date_display', '').strip()
    category = request.POST.get('category', 'Other').strip()
    image_url = request.POST.get('image_url', '').strip()
    description = request.POST.get('description', '').strip()

    if not title:
        return JsonResponse({'error': 'Event title is required'}, status=400)

    event = Event.objects.create(
        title=title,
        date_display=date_display,
        category=category,
        image_url=image_url,
        description=description,
        created_by=request.user,
    )

    # Process timeline items
    timeline_times = request.POST.getlist('timeline_time')
    timeline_activities = request.POST.getlist('timeline_activity')
    for i, (t, a) in enumerate(zip(timeline_times, timeline_activities)):
        t, a = t.strip(), a.strip()
        if t and a:
            EventTimelineItem.objects.create(event=event, time=t, activity=a, order=i)

    return JsonResponse({'success': True, 'message': f'Event "{title}" created successfully', 'id': event.id})


@require_POST
@login_required
def delete_event(request, event_id):
    """Delete an event."""
    event = get_object_or_404(Event, id=event_id)
    event.delete()
    return JsonResponse({'success': True, 'message': 'Event deleted'})


@require_GET
@login_required
def api_events_list(request):
    """Return all active events as JSON."""
    events_qs = Event.objects.filter(is_active=True)
    result = [{
        'id': e.id,
        'title': e.title,
        'date': e.date_display,
        'category': e.category,
        'img': e.image_url,
        'desc': e.description,
    } for e in events_qs]
    return JsonResponse({'events': result})


@require_POST
@login_required
def create_announcement(request):
    """Create an announcement from the admin panel."""
    title = request.POST.get('title', '').strip()
    body = request.POST.get('body', '').strip()
    importance = request.POST.get('importance', 'info').strip()
    display_day = request.POST.get('display_day', '').strip()
    display_month = request.POST.get('display_month', '').strip()
    action_link = request.POST.get('action_link', '').strip()
    action_label = request.POST.get('action_label', '').strip()

    if not title or not body:
        return JsonResponse({'error': 'Title and body are required'}, status=400)

    announcement = Announcement.objects.create(
        title=title,
        body=body,
        importance=importance,
        display_day=display_day,
        display_month=display_month,
        action_link=action_link,
        action_label=action_label,
        created_by=request.user,
    )

    return JsonResponse({'success': True, 'message': f'Announcement "{title}" published', 'id': announcement.id})


@require_POST
@login_required
def delete_announcement(request, announcement_id):
    """Delete an announcement."""
    ann = get_object_or_404(Announcement, id=announcement_id)
    ann.delete()
    return JsonResponse({'success': True, 'message': 'Announcement deleted'})


@require_GET
@login_required
def api_announcements_list(request):
    """Return all active announcements as JSON."""
    anns = Announcement.objects.filter(is_active=True)
    result = [{
        'id': a.id,
        'importance': a.importance,
        'title': a.title,
        'body': a.body,
        'date': {'day': a.display_day, 'month': a.display_month},
        'posted': a.created_at.strftime('Posted: %b %d, %Y'),
        'actionLink': a.action_link,
        'actionLabel': a.action_label,
    } for a in anns]
    return JsonResponse({'announcements': result})